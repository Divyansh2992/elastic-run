#!/usr/bin/env python3
"""
Combine all dataset_template CSVs into a single Excel workbook (.xlsx)
Each CSV becomes one named sheet.
"""
import os, glob

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import csv
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import csv

SHEET_META = {
    "01_states":              "01 - States",
    "02_cities":              "02 - Cities",
    "03_hub_operators":       "03 - Hub Operators",
    "04_fleet_operators":     "04 - Fleet Operators",
    "05_routes":              "05 - Routes",
    "06_route_services":      "06 - Route Services",
    "07_fleet_vehicle_summary": "07 - Fleet Summary",
    "08_hub_slot_timelines":  "08 - Hub Timelines",
    "09_fleet_slot_timelines":"09 - Fleet Timelines",
}

HEADER_FILL   = PatternFill("solid", fgColor="0D1420")   # dark navy
HEADER_FONT   = Font(name="Calibri", bold=True, color="00D4FF", size=10)
ROW_FILL_ODD  = PatternFill("solid", fgColor="080C14")
ROW_FILL_EVEN = PatternFill("solid", fgColor="111B2A")
ROW_FONT      = Font(name="Calibri", color="E8F4FF", size=9)
BORDER_SIDE   = Side(style="thin", color="1A2535")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                        top=BORDER_SIDE, bottom=BORDER_SIDE)

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

base = os.path.dirname(os.path.abspath(__file__))
files = sorted(glob.glob(os.path.join(base, "*.csv")))

for fpath in files:
    stem = os.path.splitext(os.path.basename(fpath))[0]
    sheet_name = SHEET_META.get(stem, stem[:31])
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    with open(fpath, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    for ri, row in enumerate(rows):
        for ci, val in enumerate(row):
            cell = ws.cell(row=ri+1, column=ci+1, value=val)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            if ri == 0:
                cell.fill   = HEADER_FILL
                cell.font   = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = ROW_FILL_ODD if ri % 2 else ROW_FILL_EVEN
                cell.font = ROW_FONT

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    ws.row_dimensions[1].height = 22

out_path = os.path.join(base, "CapacityIQ_Dataset_Template.xlsx")
wb.save(out_path)
print(f"✅ Saved: {out_path}")
